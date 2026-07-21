# ============================================
# Halbleiter & KI Aktien Ranking - Streamlit-Version v7.33
# Läuft als eigenständige Web-App (PC + Smartphone), kein Colab nötig.
# Jeder Export wird dauerhaft mit Datum + Uhrzeit in ein Google Sheet geschrieben.
# Zugriff ist per Passwort geschützt (siehe Secrets).
#
# Änderungen ggü. v7.32: siehe Changelog in Halbleiter_Ranking_v7.33.py
# (fast_info, PEG-Fallback, FCF/EV-EBITDA-Bereinigung, winsorized Min-Max-
# Scoring, verschärfte Datenqualitäts-Strafe, formatierter Excel-Download)
#
# Lokal starten:   streamlit run streamlit_app.py
# Secrets (Google Service Account, Sheet-ID, Passwort) werden NICHT im Code
# hinterlegt, sondern in Streamlit Cloud unter "Manage app" -> Settings -> Secrets.
# ============================================

import io
import time
import contextlib
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
import gspread
from google.oauth2.service_account import Credentials

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Halbleiter Ranking", page_icon="📊", layout="wide")

VERSION = "v7.33 (Streamlit + Google Sheets)"

DEFAULT_AKTIEN = [
    "MU", "SNDK", "NVDA", "AMD", "AVGO", "TSM",
    "005930.KS", "000660.KS", "285A.T", "ASML",
    "AMAT", "LRCX", "KLAC"
]

NAMEN = {
    "MU": "Micron", "SNDK": "SanDisk", "NVDA": "Nvidia", "AMD": "AMD",
    "AVGO": "Broadcom", "TSM": "TSMC", "005930.KS": "Samsung",
    "000660.KS": "SK Hynix", "285A.T": "Kioxia", "ASML": "ASML",
    "AMAT": "Applied Materials", "LRCX": "Lam Research", "KLAC": "KLA"
}

GEWICHTE = {
    "Forward KGV": 0.25, "KGV": 0.05, "PEG Ratio": 0.10, "EV/EBITDA": 0.10,
    "EPS Wachstum": 0.15, "Umsatz Wachstum": 0.10, "FCF Rendite": 0.15, "Net Debt/EBITDA": 0.10
}

GSHEET_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
LOG_WORKSHEET_NAME = "Log"


# ---------------- Passwortschutz ----------------

def check_password():
    """Zeigt ein Passwortfeld, bis das korrekte Passwort aus st.secrets eingegeben wurde."""

    def password_entered():
        if st.session_state.get("password_input") == st.secrets["app"]["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password_input"]
        else:
            st.session_state["password_correct"] = False

    if st.session_state.get("password_correct"):
        return True

    st.text_input(
        "Passwort", type="password", on_change=password_entered, key="password_input"
    )
    if st.session_state.get("password_correct") is False:
        st.error("Falsches Passwort.")
    return False


# ---------------- Google Sheets ----------------

@st.cache_resource
def get_gsheet_client():
    creds = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]), scopes=GSHEET_SCOPES
    )
    return gspread.authorize(creds)


def get_log_worksheet():
    client = get_gsheet_client()
    sheet = client.open_by_key(st.secrets["gsheet"]["sheet_id"])
    try:
        return sheet.worksheet(LOG_WORKSHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        return sheet.add_worksheet(title=LOG_WORKSHEET_NAME, rows=2000, cols=30)


def append_export_to_gsheet(df):
    ws = get_log_worksheet()
    if not ws.get_all_values():
        ws.append_row(df.columns.tolist())
    rows = df.fillna("").astype(str).values.tolist()
    ws.append_rows(rows, value_input_option="USER_ENTERED")


def get_last_export_from_gsheet():
    try:
        ws = get_log_worksheet()
        values = ws.get_all_values()
        if len(values) > 1:
            last = values[-1]
            return f"{last[0]} {last[1]}"  # Datum, Uhrzeit
    except Exception:
        pass
    return "Noch kein Export"


# ---------------- Excel-Formatierung ----------------

def format_excel_sheet(ws, df):
    """Rahmen, Freeze/Filter, Prozentformat, rechtsbündige Kennzahlen und
    farbcodierte Bewertungsspalte auf ein bereits befülltes Worksheet anwenden."""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    GRUEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    GELB = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ROT = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    header = [c.value for c in ws[1]]
    def col_idx(name):
        return header.index(name) + 1 if name in header else None

    prozent_spalten = ["EPS Wachstum", "Umsatz Wachstum", "FCF Rendite"]
    rechts_spalten = ["KGV", "Forward KGV", "PEG Ratio", "EV/EBITDA"]
    bewertung_idx = col_idx("Bewertung")

    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
        for name in prozent_spalten:
            idx = col_idx(name)
            if idx:
                row[idx - 1].number_format = '0.00"%"'
        for name in rechts_spalten:
            idx = col_idx(name)
            if idx:
                row[idx - 1].alignment = Alignment(horizontal="right")
        if bewertung_idx:
            val = row[bewertung_idx - 1].value or ""
            if "attraktiv" in val:
                row[bewertung_idx - 1].fill = GRUEN
            elif "fair" in val:
                row[bewertung_idx - 1].fill = GELB
            elif "teuer" in val:
                row[bewertung_idx - 1].fill = ROT

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def build_excel_bytes(df):
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    format_excel_sheet(ws, df)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------- Kursdaten & Ranking ----------------

def _fast_info_get(fast, key_camel, key_snake, default=None):
    """Robuster Zugriff auf fast_info: die Attribut-Namen unterscheiden sich
    zwischen yfinance-Versionen (camelCase-Keys vs. snake_case-Attribute).
    Nicht live getestet (kein Netzwerkzugriff in dieser Umgebung) -
    vor produktivem Einsatz einmal gegen die installierte yfinance-Version prüfen."""
    for accessor in (
        lambda: fast[key_camel],
        lambda: getattr(fast, key_snake),
        lambda: fast.get(key_camel),
    ):
        try:
            val = accessor()
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                return val
        except Exception:
            continue
    return default


def get_yahoo_data(symbol, is_pflicht=False):
    max_retries = 10 if is_pflicht else 3
    for _ in range(max_retries):
        try:
            f = io.StringIO()
            with contextlib.redirect_stderr(f), contextlib.redirect_stdout(f):
                ticker = yf.Ticker(symbol)
                fast = ticker.fast_info

            kurs = _fast_info_get(fast, "lastPrice", "last_price")
            marketcap = _fast_info_get(fast, "marketCap", "market_cap")

            if kurs is None or pd.isna(kurs):
                with contextlib.redirect_stderr(f), contextlib.redirect_stdout(f):
                    hist = ticker.history(period="1d")
                kurs = hist["Close"].iloc[-1] if not hist.empty else None

            try:
                with contextlib.redirect_stderr(f), contextlib.redirect_stdout(f):
                    info = ticker.info or {}
            except Exception:
                info = {}

            if kurs is None or pd.isna(kurs):
                kurs = info.get("currentPrice") or info.get("regularMarketPrice")
            if kurs is None or pd.isna(kurs):
                time.sleep(2)
                continue

            if not marketcap:
                marketcap = info.get("marketCap")

            kgv = info.get("trailingPE")
            if kgv is None and kurs and info.get("trailingEps") and info.get("trailingEps") != 0:
                kgv = kurs / info.get("trailingEps")

            forward_pe = info.get("forwardPE")
            growth = info.get("earningsGrowth")

            peg = info.get("pegRatio")
            if peg is None and forward_pe and growth and growth > 0:
                peg = forward_pe / (growth * 100)

            fcf = info.get("freeCashflow")
            fcf_rendite = np.nan
            if fcf is not None and marketcap:
                fcf_rendite = (fcf / marketcap) if fcf > 0 else np.nan

            debt = info.get("totalDebt"); cash = info.get("totalCash"); ebitda = info.get("ebitda")
            net_debt_ebitda = ((debt - cash) / ebitda) if (debt is not None and cash is not None and ebitda and ebitda > 0) else None

            return {
                "Ticker": symbol, "Name": info.get("shortName") or st.session_state.namen.get(symbol, symbol),
                "Kurs": kurs, "KGV": kgv, "Forward KGV": forward_pe,
                "PEG Ratio": peg, "EV/EBITDA": info.get("enterpriseToEbitda"),
                "EPS Wachstum": growth, "Umsatz Wachstum": info.get("revenueGrowth"),
                "FCF Rendite": fcf_rendite, "Net Debt/EBITDA": net_debt_ebitda
            }
        except Exception:
            time.sleep(3 if is_pflicht else 1.5)
    return "PFLICHT_FEHLER" if is_pflicht else None


def run_ranking(aktien_liste, status_placeholder, progress_bar):
    daten, fehlgeschlagen = [], []

    for i, symbol in enumerate(aktien_liste):
        is_pflicht = (symbol == "SNDK")
        status_placeholder.write(f"Lade {symbol} …")
        data = get_yahoo_data(symbol, is_pflicht)

        if data == "PFLICHT_FEHLER":
            st.error("⚠️ SERVER DOWN — SNDK konnte nach 10 Versuchen nicht geladen werden. Bitte später nochmals versuchen.")
            return None, fehlgeschlagen
        if data:
            daten.append(data)
        else:
            fehlgeschlagen.append(symbol)
        progress_bar.progress((i + 1) / len(aktien_liste))

    if len(daten) < 5:
        st.error("❌ Zu wenige Daten. Bitte später nochmal probieren.")
        return None, fehlgeschlagen

    df = pd.DataFrame(daten)
    for c in df.columns:
        if c not in ["Ticker", "Name"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    for col in ["Forward KGV", "KGV", "PEG Ratio"]:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: np.nan if (pd.notna(x) and x <= 0) else x)

    if "EV/EBITDA" in df.columns:
        df["EV/EBITDA"] = df["EV/EBITDA"].apply(
            lambda x: np.nan if (pd.notna(x) and (x <= 0 or x > 100)) else x
        )

    def niedrig_besser(x, q_low=0.05, q_high=0.95):
        lo, hi = x.quantile(q_low), x.quantile(q_high)
        if pd.isna(lo) or pd.isna(hi) or hi == lo:
            return pd.Series(50.0, index=x.index)
        clipped = x.clip(lo, hi)
        return (hi - clipped) / (hi - lo) * 100

    def hoch_besser(x, q_low=0.05, q_high=0.95):
        lo, hi = x.quantile(q_low), x.quantile(q_high)
        if pd.isna(lo) or pd.isna(hi) or hi == lo:
            return pd.Series(50.0, index=x.index)
        clipped = x.clip(lo, hi)
        return (clipped - lo) / (hi - lo) * 100

    niedrig = ["Forward KGV", "KGV", "PEG Ratio", "EV/EBITDA", "Net Debt/EBITDA"]

    score = pd.Series(0.0, index=df.index)
    gewicht_vorhanden = pd.Series(0.0, index=df.index)
    for kennzahl, gewicht in GEWICHTE.items():
        if df[kennzahl].notna().sum() < 2:
            continue
        einzel_score = niedrig_besser(df[kennzahl]) if kennzahl in niedrig else hoch_besser(df[kennzahl])
        vorhanden = df[kennzahl].notna()
        score += einzel_score.where(vorhanden, 0) * gewicht
        gewicht_vorhanden += vorhanden * gewicht

    df["Gesamtscore"] = (score / gewicht_vorhanden.replace(0, np.nan)).round(1)
    datenqualitaet = pd.Series(0.0, index=df.index)
    for k, w in GEWICHTE.items():
        datenqualitaet += df[k].notna() * w
    df["Datenqualität"] = (datenqualitaet * 100).round(0)
    df["Adjustierter Score"] = (df["Gesamtscore"] * (0.5 + 0.5 * (df["Datenqualität"] / 100))).round(1)

    df["Bewertung"] = df["Adjustierter Score"].apply(lambda x: "🟢 attraktiv" if x >= 75 else "🟡 fair" if x >= 50 else "🔴 teuer")
    df["Kurs"] = df["Kurs"].round(2)
    for c in ["KGV", "Forward KGV", "EV/EBITDA"]:
        df[c] = df[c].round(1)
    df["PEG Ratio"] = df["PEG Ratio"].round(2)
    for c in ["EPS Wachstum", "Umsatz Wachstum", "FCF Rendite"]:
        df[c] = (df[c] * 100).round(2)
    df = df.sort_values("Adjustierter Score", ascending=False)

    jetzt = datetime.now()
    df.insert(0, "Uhrzeit", jetzt.strftime("%H:%M:%S"))
    df.insert(0, "Datum", jetzt.strftime("%Y-%m-%d"))

    status_placeholder.empty()
    progress_bar.empty()
    return df, fehlgeschlagen


# ================= UI =================

if not check_password():
    st.stop()

if "aktien" not in st.session_state:
    st.session_state.aktien = DEFAULT_AKTIEN.copy()
if "namen" not in st.session_state:
    st.session_state.namen = NAMEN.copy()
if "last_export" not in st.session_state:
    st.session_state.last_export = get_last_export_from_gsheet()
if "df_result" not in st.session_state:
    st.session_state.df_result = None

st.title("📊 Halbleiter & KI Aktien Ranking")
st.caption(f"{VERSION} · Letzter Export: {st.session_state.last_export}")

col1, col2 = st.columns([3, 1])
with col1:
    new_ticker = st.text_input(
        "Ticker hinzufügen", placeholder="z.B. GOOGL oder TSM", label_visibility="collapsed"
    )
with col2:
    if st.button("➕ Hinzufügen", use_container_width=True):
        t = new_ticker.upper().strip()
        if not t:
            pass
        elif t in st.session_state.aktien:
            st.warning("Der Wert ist schon vorhanden.")
        else:
            data = get_yahoo_data(t)
            if data:
                st.session_state.aktien.append(t)
                st.session_state.namen[t] = data["Name"]
                st.success(f"{t} ({data['Name']}) erfolgreich hinzugefügt.")
            else:
                st.error("Diesen Wert gibt es nicht oder der Server ist down.")

st.write(f"**Aktuelle Liste ({len(st.session_state.aktien)}):** " + ", ".join(st.session_state.aktien))

if st.button("🚀 Ranking starten", type="primary"):
    status = st.empty()
    progress_bar = st.progress(0)
    df, fehlgeschlagen = run_ranking(st.session_state.aktien, status, progress_bar)
    if df is not None:
        st.session_state.df_result = df
        if fehlgeschlagen:
            st.warning(f"Übersprungen: {', '.join(fehlgeschlagen)}")
        try:
            append_export_to_gsheet(df)
            st.session_state.last_export = f"{df['Datum'].iloc[0]} {df['Uhrzeit'].iloc[0]}"
            st.success("✅ Fertig – im Google Sheet gespeichert")
        except Exception as e:
            st.warning(f"Ranking fertig, aber Speichern im Google Sheet ist fehlgeschlagen: {e}")

if st.session_state.df_result is not None:
    st.dataframe(st.session_state.df_result, use_container_width=True, hide_index=True)

    st.download_button(
        "⬇️ Als Excel herunterladen",
        data=build_excel_bytes(st.session_state.df_result),
        file_name=f"Halbleiter_Ranking_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
